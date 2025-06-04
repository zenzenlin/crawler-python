import time
import random
import requests
import openpyxl
import json



# ----------------------------------------------------------------
# 取得公司編號
def search_company_encode(page:int = 1, filter_params:dict = None) -> list[str]:
  print('Step 2. search_company_encode ')
  # 公司編號列表
  encoded_no = []
  # 構建 URL
  url = 'https://www.104.com.tw/company/ajax/list'
  # 篩選條件
  query = f'?jobsource=cs_custlist&mode=s&pageSize=18'
  # header 資訊
  headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.92 Safari/537.36',
    'Referer': 'https://www.104.com.tw/company/search/',
  }

  # 加上篩選參數，先轉換為 URL 參數字串格式
  if filter_params:
    query += ''.join([f'&{key}={value}' for key, value, in filter_params.items()])

  params = f'{query}&page={page}'
  print('params ', params)

  try:
    r = requests.get(url, params=params, headers=headers)
    # 如果請求失敗，將會拋出異常
    r.raise_for_status()

    data = r.json()

    # 迴圈取得公司編號
    encoded_no = [company_data['encodedCustNo'] for company_data in data['data']]

    # 無資料或已達最後頁
    if (page == data['metadata']['pagination']['total']) or (data['metadata']['pagination']['total'] == 0):
      print('no data or last page')

  except requests.exceptions.RequestException as e:
    # 請求失敗時的處理，將異常實例賦值給變數 e
    print('請求失敗:', e)

  return encoded_no


# ----------------------------------------------------------------
# 取得公司資訊，寫入 Excel
def search_companies(max_page:int = 10, filter_params:dict = None, area_name:str = '台北市'):
  # 公司列表
  company_list = []
  # 公司編號列表
  encodes = []
  # 頁數
  page = 1

  # 小於等於 max_page 時，執行迴圈
  while page <= max_page:
    print('Step 1. search_companies gets encodes ', page)
    print('Step 1. key ', filter_params, area_name)
    encodes = search_company_encode(page=page, filter_params=filter_params)

    # 如果沒有資料，退出目前的迴圈並開始下一個迴圈
    if not encodes:
      print('No data for this page. Moving to the next.')
      break

    # 迴圈取得公司資訊
    for encode in encodes:
      print('Step 3. search_companies gets each company info ', encode)
      # 構建 URL
      url = f'https://www.104.com.tw/company/ajax/content/{encode}?'
      # header 資訊
      headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.92 Safari/537.36',
        'Referer': 'https://www.104.com.tw/company/',
      }

      try:
        r = requests.get(url, headers=headers)
        # 如果請求失敗，將會拋出異常
        r.raise_for_status()

        data = r.json()

        # 取出指定的 key value 組合單一公司資訊
        company_info = {
          'custName': data['data']['custName'],
          'indcat': data['data']['indcat'],
          'capital': data['data']['capital'],
          'empNo': data['data']['empNo'],
          'hrName': data['data']['hrName'],
          'phone': data['data']['phone'],
          'address': data['data']['address']
        }

        # 將單一公司資訊塞回公司列表
        company_list.append(company_info)

        # 暫停 1 秒，以免對 server 造成過大壓力
        time.sleep(1)

        # 如果 公司列表頁數長度 (每頁 18 筆計) 等於 max_page，退出迴圈
        if len(company_list) // 18 == max_page:
          break

      except requests.exceptions.RequestException as e:
        # 請求失敗時的處理，將異常實例賦值給變數 e
        print('請求失敗:', e)
        continue

    # 打開現有的 Excel 檔案，若不存在則建立一個新的工作簿
    try:
      workbook = openpyxl.load_workbook('company_list.xlsx')
    except FileNotFoundError:
      workbook = openpyxl.Workbook()

    # 取得指定名稱工作表，若不存在則建立一個新的指定名稱工作表
    sheet_name = area_name
    if sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
    else:
      sheet = workbook.create_sheet(sheet_name, 0)

    # 將 column header 放入每個公司列表 key
    headers = list(company_list[0].keys())
    for col_idx, header in enumerate(headers, start=1):
      sheet.cell(row=1, column=col_idx).value = header

    # 將資料寫入工作表
    for row_idx, item in enumerate(company_list, start=2):
      for col_idx, key in enumerate(headers, start=1):
        try:
          sheet.cell(row=row_idx, column=col_idx).value = item.get(key, '')
        except openpyxl.utils.exceptions.IllegalCharacterError as e:
          # 資料中存有預期外的字元，跳過此項
          print(f"Illegal character encountered: {e}. Skipping this item.")
          continue

    # 儲存檔案
    workbook.save('company_list.xlsx')

    page += 1



# ----------------------------------------------------------------
# 區域名稱
area_name = '台北市中正區'
# 搜尋條件
filter_params = {
  'area': '6001001001',  # (地區) 台北市
  'jobsource': 'cs_custlist',
  'mode': 's',
  'pageSize': '18',
}

# 迴圈處理 JSON 中的每個區域
with open('area.json', 'r') as f:
  data = json.load(f)

  for i in data:
    for j in i['n']:
      area_name = j['des']
      filter_params = {
        'area': j['no']
      }
      search_companies(max_page=100, filter_params=filter_params, area_name=area_name)
